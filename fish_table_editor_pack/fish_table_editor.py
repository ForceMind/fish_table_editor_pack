#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import socket
from datetime import datetime
from dataclasses import dataclass
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, List, Tuple
from urllib.parse import urlparse
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
TABLE_DIR = APP_DIR.parent
WEB_DIR = APP_DIR / "table_editor_web"
PRESET_DIR = APP_DIR / "script_presets"
SCRIPT_FILE = TABLE_DIR / "Script&.xlsx"
FILE_MAP = {
    "arena": TABLE_DIR / "Arena&.xlsx",
    "fish": TABLE_DIR / "Fish&.xlsx",
    "group": TABLE_DIR / "Group&.xlsx",
    "route": TABLE_DIR / "Route&.xlsx",
    "script": TABLE_DIR / "Script&.xlsx",
}
OUTPUT_FILE_RE = re.compile(r"^[\w\-.&\u4e00-\u9fa5 ]+\.xlsx$", re.IGNORECASE)


def parse_id_list(value: Any) -> List[int]:
    if value is None:
        return []
    if isinstance(value, int):
        return [value] if value > 0 else []
    if isinstance(value, float):
        n = int(value)
        return [n] if n > 0 else []
    if isinstance(value, str):
        out: List[int] = []
        for part in value.split(","):
            token = part.strip()
            if not token:
                continue
            try:
                n = int(float(token))
                if n > 0:
                    out.append(n)
            except ValueError:
                continue
        return out
    if isinstance(value, list):
        out: List[int] = []
        for item in value:
            out.extend(parse_id_list(item))
        return out
    return []


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def clamp_float(value: Any, min_value: float, max_value: float, default: float) -> float:
    try:
        n = float(value)
    except (TypeError, ValueError):
        return default
    return max(min_value, min(max_value, n))


def sanitize_llm_config(config: Any) -> Dict[str, Any]:
    data = config if isinstance(config, dict) else {}
    base_url = str(data.get("baseUrl") or os.environ.get("OPENAI_BASE_URL") or "https://api.openai.com/v1").strip()
    if not base_url:
        base_url = "https://api.openai.com/v1"
    while base_url.endswith("/"):
        base_url = base_url[:-1]
    model = str(data.get("model") or os.environ.get("OPENAI_MODEL") or "gpt-4.1-mini").strip() or "gpt-4.1-mini"
    api_key = str(data.get("apiKey") or os.environ.get("OPENAI_API_KEY") or "").strip()
    temperature = clamp_float(data.get("temperature", os.environ.get("OPENAI_TEMPERATURE", 0.2)), 0.0, 1.0, 0.2)
    timeout_sec = max(20, to_int(data.get("timeoutSec", os.environ.get("OPENAI_TIMEOUT_SEC", 90)), 90))
    max_tokens = max(512, to_int(data.get("maxTokens", os.environ.get("OPENAI_MAX_TOKENS", 2048)), 2048))
    return {
        "baseUrl": base_url,
        "model": model,
        "apiKey": api_key,
        "temperature": temperature,
        "timeoutSec": timeout_sec,
        "maxTokens": max_tokens,
    }


def sanitize_output_name(name: str) -> str:
    safe = os.path.basename((name or "").strip())
    if safe and not safe.lower().endswith(".xlsx"):
        safe += ".xlsx"
    return safe


def ensure_unique_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix or ".xlsx"
    index = 1
    while True:
        candidate = path.with_name(f"{stem}_{index:02d}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def build_output_path(output_name: str) -> Path:
    safe = sanitize_output_name(output_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = SCRIPT_FILE.name.lower()
    if not safe or safe.lower() == base_name:
        candidate = TABLE_DIR / f"Script.generated.{timestamp}.xlsx"
    else:
        stem = Path(safe).stem
        candidate = TABLE_DIR / f"{stem}.{timestamp}.xlsx"
    return ensure_unique_output_path(candidate)


def ensure_preset_dir() -> None:
    PRESET_DIR.mkdir(parents=True, exist_ok=True)


def sanitize_preset_name(name: str) -> str:
    preset = (name or "").strip()
    if not preset:
        raise ValueError("preset name is required")
    if preset.lower().endswith(".json"):
        preset = preset[:-5]
    if not re.match(r"^[\w\-.&\u4e00-\u9fa5 ]+$", preset):
        raise ValueError("preset name illegal")
    return preset


def normalize_scripts_payload(scripts: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    normalized: List[Dict[str, Any]] = []
    warnings: List[str] = []
    for idx, row in enumerate(scripts):
        script_id = to_int(row.get("scriptId"))
        group_ids = parse_id_list(row.get("groupIds"))
        arena_ids = parse_id_list(row.get("arenaIds"))
        gap_time = to_int(row.get("gapTimeMs"), 0)
        row_type = to_int(row.get("type"), 1)
        if script_id <= 0:
            warnings.append(f"row {idx + 1}: scriptId invalid")
            continue
        if not group_ids:
            warnings.append(f"row {idx + 1}: groupIds empty")
            continue
        if not arena_ids:
            warnings.append(f"row {idx + 1}: arenaIds empty")
            continue
        normalized.append(
            {
                "scriptId": script_id,
                "groupIds": group_ids,
                "gapTimeMs": max(0, gap_time),
                "arenaIds": arena_ids,
                "type": row_type,
            }
        )
    return normalized, warnings


def preset_file_by_name(name: str) -> Path:
    ensure_preset_dir()
    safe = sanitize_preset_name(name)
    return PRESET_DIR / f"{safe}.json"


def list_presets() -> List[Dict[str, Any]]:
    ensure_preset_dir()
    out: List[Dict[str, Any]] = []
    for file_path in sorted(PRESET_DIR.glob("*.json"), key=lambda p: p.name.lower()):
        row_count = 0
        created_at = ""
        try:
            payload = json.loads(file_path.read_text(encoding="utf-8"))
            scripts = payload.get("scripts", [])
            if isinstance(scripts, list):
                row_count = len(scripts)
            created_at = str(payload.get("createdAt") or "")
        except (OSError, json.JSONDecodeError):
            pass
        out.append(
            {
                "name": file_path.stem,
                "file": file_path.name,
                "rows": row_count,
                "createdAt": created_at,
            }
        )
    return out


def save_preset(name: str, scripts: List[Dict[str, Any]], meta: Dict[str, Any] | None = None) -> Dict[str, Any]:
    file_path = preset_file_by_name(name)
    normalized, warnings = normalize_scripts_payload(scripts)
    payload = {
        "name": file_path.stem,
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "rows": len(normalized),
        "meta": meta or {},
        "scripts": normalized,
    }
    file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "name": file_path.stem,
        "file": file_path.name,
        "rows": len(normalized),
        "warnings": warnings,
    }


def load_preset(name: str) -> Dict[str, Any]:
    file_path = preset_file_by_name(name)
    if not file_path.exists() or not file_path.is_file():
        raise FileNotFoundError("preset not found")
    payload = json.loads(file_path.read_text(encoding="utf-8"))
    scripts = payload.get("scripts", [])
    if not isinstance(scripts, list):
        raise ValueError("invalid preset: scripts must be array")
    normalized, warnings = normalize_scripts_payload(scripts)
    return {
        "name": file_path.stem,
        "file": file_path.name,
        "rows": len(normalized),
        "meta": payload.get("meta", {}),
        "scripts": normalized,
        "warnings": warnings,
    }


def delete_preset(name: str) -> Dict[str, Any]:
    file_path = preset_file_by_name(name)
    if not file_path.exists() or not file_path.is_file():
        raise FileNotFoundError("preset not found")
    file_path.unlink()
    return {"name": file_path.stem, "file": file_path.name}


def load_sheet_rows(file_path: Path) -> Tuple[str, List[List[Any]]]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return ws.title, rows


def build_keys(raw_keys: List[Any]) -> List[str]:
    keys: List[str] = []
    seen: Dict[str, int] = {}
    for item in raw_keys:
        key = str(item).strip() if item is not None else ""
        if not key:
            keys.append("")
            continue
        count = seen.get(key, 0) + 1
        seen[key] = count
        keys.append(key if count == 1 else f"{key}_{count}")
    return keys


def read_table_records(file_path: Path) -> List[Dict[str, Any]]:
    _, rows = load_sheet_rows(file_path)
    if len(rows) < 2:
        return []
    keys = build_keys(rows[1])
    out: List[Dict[str, Any]] = []
    for row in rows[2:]:
        if not any(cell not in (None, "") for cell in row[: len(keys)]):
            continue
        rec: Dict[str, Any] = {}
        for idx, key in enumerate(keys):
            if not key:
                continue
            rec[key] = row[idx] if idx < len(row) else None
        out.append(rec)
    return out


def format_payout(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def read_fish_records(file_path: Path) -> List[Dict[str, Any]]:
    records = read_table_records(file_path)
    out: List[Dict[str, Any]] = []
    for rec in records:
        fish_id = to_int(rec.get("Id"))
        if fish_id <= 0:
            continue
        en_name = str(rec.get("Name") or f"Fish-{fish_id}")
        cn_name = str(rec.get("#Note") or rec.get("Note") or "").strip()
        cn_ext = str(rec.get("#Note_2") or rec.get("Note_2") or "").strip()
        fish_type = to_int(rec.get("FishType"), 0)
        payout = to_float(rec.get("Payout"), 0.0)
        if not cn_name:
            cn_name = en_name
        cn_full = cn_name if not cn_ext else f"{cn_name}·{cn_ext}"
        out.append(
            {
                "id": fish_id,
                "name": en_name,
                "cnName": cn_name,
                "cnFullName": cn_full,
                "fishType": fish_type,
                "payout": round(payout, 2),
                "payoutText": format_payout(payout),
                "label": f"{cn_full}({format_payout(payout)}倍)",
            }
        )
    out.sort(key=lambda x: x["id"])
    return out


@dataclass
class CacheData:
    arenas: List[Dict[str, Any]]
    fish: List[Dict[str, Any]]
    groups: List[Dict[str, Any]]
    routes: List[Dict[str, Any]]
    scripts: List[Dict[str, Any]]


def load_all_data() -> CacheData:
    arena_rows = read_table_records(FILE_MAP["arena"])
    fish_rows = read_fish_records(FILE_MAP["fish"])
    group_rows = read_table_records(FILE_MAP["group"])
    route_rows = read_table_records(FILE_MAP["route"])
    script_rows = read_table_records(FILE_MAP["script"])

    fish_by_id = {item["id"]: item for item in fish_rows}

    route_by_id: Dict[int, Dict[str, Any]] = {}
    routes: List[Dict[str, Any]] = []
    for rec in route_rows:
        route_id = to_int(rec.get("Id"))
        if route_id <= 0:
            continue
        route = {
            "id": route_id,
            "routeId": to_int(rec.get("RouteId"), route_id),
            "time": round(to_float(rec.get("Time"), 0.0), 2),
            "animation": str(rec.get("Animation") or ""),
        }
        route_by_id[route_id] = route
        routes.append(route)
    routes.sort(key=lambda x: x["id"])

    arenas: List[Dict[str, Any]] = []
    for rec in arena_rows:
        arena_id = to_int(rec.get("Id"))
        if arena_id <= 0:
            continue
        arenas.append(
            {
                "id": arena_id,
                "name": str(rec.get("Name") or f"Arena-{arena_id}"),
                "gunCount": to_int(rec.get("GunCount"), 0),
                "scene": to_int(rec.get("Scene"), 0),
            }
        )
    arenas.sort(key=lambda x: x["id"])

    groups: List[Dict[str, Any]] = []
    for rec in group_rows:
        group_id = to_int(rec.get("GroupId"))
        if group_id <= 0:
            continue
        fish_ids = parse_id_list(rec.get("Fish"))
        route_ids = parse_id_list(rec.get("RouteId"))
        fish_items = [
            fish_by_id.get(
                fish_id,
                {
                    "id": fish_id,
                    "name": f"Fish-{fish_id}",
                    "cnName": f"Fish-{fish_id}",
                    "cnFullName": f"Fish-{fish_id}",
                    "fishType": 0,
                    "payout": 0.0,
                    "payoutText": "0",
                    "label": f"Fish-{fish_id}(0倍)",
                },
            )
            for fish_id in fish_ids
        ]
        payouts = [item["payout"] for item in fish_items if item["payout"] > 0]
        route_times = [route_by_id.get(route_id, {}).get("time", 0.0) for route_id in route_ids]
        route_times = [x for x in route_times if x > 0]
        avg_payout = round(sum(payouts) / len(payouts), 2) if payouts else 0.0
        avg_route_time = round(sum(route_times) / len(route_times), 2) if route_times else 0.0

        composition_count: Dict[int, int] = {}
        for fish_id in fish_ids:
            composition_count[fish_id] = composition_count.get(fish_id, 0) + 1
        composition: List[Dict[str, Any]] = []
        for fish_id, count in composition_count.items():
            fish_info = fish_by_id.get(fish_id)
            if fish_info is None:
                composition.append(
                    {
                        "fishId": fish_id,
                        "count": count,
                        "cnName": f"Fish-{fish_id}",
                        "name": f"Fish-{fish_id}",
                        "payout": 0.0,
                        "payoutText": "0",
                    }
                )
            else:
                composition.append(
                    {
                        "fishId": fish_id,
                        "count": count,
                        "cnName": fish_info["cnFullName"],
                        "name": fish_info["name"],
                        "payout": fish_info["payout"],
                        "payoutText": fish_info["payoutText"],
                    }
                )
        composition.sort(key=lambda x: x["fishId"])

        groups.append(
            {
                "id": group_id,
                "fishIds": fish_ids,
                "fishList": fish_items,
                "fishCnNames": [item["cnFullName"] for item in fish_items],
                "fishNames": [item["name"] for item in fish_items],
                "fishLabels": [item["label"] for item in fish_items],
                "composition": composition,
                "routeIds": route_ids,
                "routeTimes": route_times,
                "gapTime": to_int(rec.get("GapTime"), 0),
                "type": to_int(rec.get("Type"), 0),
                "avgPayout": avg_payout,
                "avgRouteTime": avg_route_time,
                "hasBoss": any(item["fishType"] >= 100 for item in fish_items),
                "hasSkill": any(item["fishType"] in (2, 3) for item in fish_items),
            }
        )
    groups.sort(key=lambda x: x["id"])

    scripts: List[Dict[str, Any]] = []
    for rec in script_rows:
        script_id = to_int(rec.get("ScriptId"))
        if script_id <= 0:
            continue
        group_ids = parse_id_list(rec.get("Group"))
        arena_ids = parse_id_list(rec.get("Arena"))
        if not group_ids or not arena_ids:
            continue
        scripts.append(
            {
                "scriptId": script_id,
                "groupIds": group_ids,
                "gapTimeMs": to_int(rec.get("GapTime"), 0),
                "arenaIds": arena_ids,
                "type": to_int(rec.get("Type"), 1),
            }
        )

    return CacheData(arenas=arenas, fish=fish_rows, groups=groups, routes=routes, scripts=scripts)


def group_fish_count(group: Dict[str, Any]) -> int:
    fish_list = group.get("fishList", [])
    if isinstance(fish_list, list) and fish_list:
        return len(fish_list)
    fish_ids = parse_id_list(group.get("fishIds"))
    if fish_ids:
        return len(fish_ids)
    comp = group.get("composition", [])
    if isinstance(comp, list) and comp:
        return sum(max(0, to_int(x.get("count"), 0)) for x in comp if isinstance(x, dict))
    return 0


def build_llm_generation_context(cache: CacheData, min_per_arena: int) -> Dict[str, Any]:
    group_by_id = {g["id"]: g for g in cache.groups}
    arena_to_groups: Dict[int, set[int]] = {}
    arena_to_scripts: Dict[int, List[Dict[str, Any]]] = {}
    for row in cache.scripts:
        for arena_id in row.get("arenaIds", []):
            if arena_id <= 0:
                continue
            if arena_id not in arena_to_groups:
                arena_to_groups[arena_id] = set()
            if arena_id not in arena_to_scripts:
                arena_to_scripts[arena_id] = []
            for gid in row.get("groupIds", []):
                if gid > 0:
                    arena_to_groups[arena_id].add(gid)
            arena_to_scripts[arena_id].append(
                {
                    "scriptId": to_int(row.get("scriptId"), 0),
                    "gapTimeMs": max(0, to_int(row.get("gapTimeMs"), 0)),
                    "groupIds": parse_id_list(row.get("groupIds")),
                }
            )

    arena_payload: List[Dict[str, Any]] = []
    for arena in cache.arenas:
        arena_id = arena["id"]
        configured_ids = sorted(arena_to_groups.get(arena_id, set()))
        groups: List[Dict[str, Any]] = []
        for gid in configured_ids:
            g = group_by_id.get(gid)
            if not g:
                continue
            fish_count = group_fish_count(g)
            if fish_count <= 0:
                continue
            groups.append(
                {
                    "id": g["id"],
                    "avgPayout": float(g.get("avgPayout", 0.0) or 0.0),
                    "fishCount": fish_count,
                    "hasBoss": bool(g.get("hasBoss")),
                    "hasSkill": bool(g.get("hasSkill")),
                    "routeCount": len(parse_id_list(g.get("routeIds"))),
                }
            )
        if not groups:
            continue
        arena_payload.append(
            {
                "arenaId": arena_id,
                "name": arena.get("name", f"Arena-{arena_id}"),
                "minScripts": min_per_arena,
                "groups": groups,
                # Keep sample scripts short to reduce context size and timeout risk.
                "existingScripts": sorted(arena_to_scripts.get(arena_id, []), key=lambda x: x["scriptId"])[:8],
            }
        )

    start_script_id = max((to_int(x.get("scriptId"), 0) for x in cache.scripts), default=0) + 1
    return {
        "startScriptId": start_script_id,
        "minPerArena": min_per_arena,
        "arenas": arena_payload,
    }


def _content_to_text(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        for key in ("text", "content", "value", "output_text"):
            v = value.get(key)
            if isinstance(v, str) and v.strip():
                return v
        text_obj = value.get("text")
        if isinstance(text_obj, dict):
            for key in ("value", "content", "text"):
                v = text_obj.get(key)
                if isinstance(v, str) and v.strip():
                    return v
        return ""
    if isinstance(value, list):
        parts = [_content_to_text(x) for x in value]
        return "".join(x for x in parts if x)
    return ""


def _extract_completion_candidate(payload: Dict[str, Any]) -> Tuple[Any, str]:
    if not isinstance(payload, dict):
        return "", "payload-not-dict"
    diag: List[str] = []
    choices = payload.get("choices", [])
    if isinstance(choices, list) and choices:
        c0 = choices[0] if isinstance(choices[0], dict) else {}
        msg = c0.get("message", {})
        if isinstance(msg, dict):
            parsed = msg.get("parsed")
            if isinstance(parsed, dict):
                return parsed, "message.parsed"
            content = msg.get("content")
            if isinstance(content, (dict, list, str)):
                if isinstance(content, dict):
                    return content, "message.content.dict"
                text = _content_to_text(content)
                if text.strip():
                    return text, "message.content"
            tool_calls = msg.get("tool_calls", [])
            if isinstance(tool_calls, list) and tool_calls:
                tc0 = tool_calls[0] if isinstance(tool_calls[0], dict) else {}
                func = tc0.get("function", {})
                if isinstance(func, dict):
                    args = func.get("arguments")
                    if isinstance(args, str) and args.strip():
                        return args, "message.tool_calls.arguments"
            refusal = msg.get("refusal")
            if isinstance(refusal, str) and refusal.strip():
                diag.append(f"refusal:{refusal[:140]}")
        text = c0.get("text")
        if isinstance(text, str) and text.strip():
            return text, "choice.text"
    out_text = payload.get("output_text")
    if isinstance(out_text, str) and out_text.strip():
        return out_text, "output_text"
    top_content = payload.get("content")
    if isinstance(top_content, (str, list, dict)):
        text = _content_to_text(top_content)
        if text.strip():
            return text, "top.content"
    return "", ";".join(diag) or "no-supported-content-field"


def _extract_finish_reason(payload: Dict[str, Any]) -> str:
    if not isinstance(payload, dict):
        return ""
    choices = payload.get("choices", [])
    if isinstance(choices, list) and choices:
        c0 = choices[0] if isinstance(choices[0], dict) else {}
        reason = c0.get("finish_reason")
        if isinstance(reason, str):
            return reason
    return ""


def _strip_json_fence(text: str) -> str:
    t = (text or "").strip()
    if not t:
        return t
    if t.startswith("```"):
        start = t.find("{")
        end = t.rfind("}")
        if start >= 0 and end > start:
            return t[start : end + 1]
    return t


def _extract_first_json_object(text: str) -> str:
    s = text or ""
    start = s.find("{")
    if start < 0:
        return ""
    depth = 0
    in_str = False
    escape = False
    for i in range(start, len(s)):
        ch = s[i]
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return s[start : i + 1]
    return ""


def _push_debug(debug_trace: List[Dict[str, Any]], title: str, content: Any) -> None:
    debug_trace.append({"title": title, "content": content})


def _post_chat_completion(
    endpoint: str,
    api_key: str,
    body: Dict[str, Any],
    timeout_sec: int,
    debug_trace: List[Dict[str, Any]],
    attempt_tag: str,
) -> Dict[str, Any]:
    _push_debug(debug_trace, f"{attempt_tag} request body", body)
    req = Request(
        endpoint,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(req, timeout=timeout_sec) as resp:
            raw_text = resp.read().decode("utf-8", errors="replace")
            _push_debug(debug_trace, f"{attempt_tag} raw response text", raw_text)
            payload = json.loads(raw_text)
            _push_debug(debug_trace, f"{attempt_tag} parsed response payload", payload)
            return payload
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        _push_debug(debug_trace, f"{attempt_tag} http error body", detail)
        raise RuntimeError(f"大模型接口错误: HTTP {exc.code} {detail}") from exc
    except (TimeoutError, socket.timeout) as exc:
        _push_debug(debug_trace, f"{attempt_tag} timeout", str(exc))
        raise RuntimeError(f"大模型请求超时（{timeout_sec}s）: {exc}") from exc
    except URLError as exc:
        raise RuntimeError(f"大模型接口不可达: {exc}") from exc


def _debug_print_json(title: str, obj: Any) -> None:
    try:
        print(f"[llm-debug] {title}")
        print(json.dumps(obj, ensure_ascii=False, indent=2))
    except Exception:  # pylint: disable=broad-except
        print(f"[llm-debug] {title} (json print failed)")
        print(str(obj))


def _debug_print_text(title: str, text: str) -> None:
    print(f"[llm-debug] {title}")
    print(text if isinstance(text, str) else str(text))


def call_openai_for_scripts(
    context: Dict[str, Any],
    llm_config: Dict[str, Any],
    debug_trace: List[Dict[str, Any]],
    arena_tag: str = "",
) -> Tuple[List[Dict[str, Any]], str, str]:
    api_key = str(llm_config.get("apiKey") or "").strip()
    if not api_key:
        raise RuntimeError("未设置 OPENAI_API_KEY，无法使用大模型生成")
    model = str(llm_config.get("model") or "gpt-4.1-mini").strip() or "gpt-4.1-mini"
    timeout_sec = max(20, to_int(llm_config.get("timeoutSec"), 90))
    temperature = clamp_float(llm_config.get("temperature"), 0.0, 1.0, 0.2)
    max_tokens = max(512, to_int(llm_config.get("maxTokens"), 2048))
    base_url = str(llm_config.get("baseUrl") or "https://api.openai.com/v1").strip()
    endpoint = f"{base_url}/chat/completions"

    system_prompt = (
        "你是资深捕鱼策划，需要输出可落地的脚本编排JSON。"
        "目标：让对局有节奏、有波峰、有缓冲，且满足工程约束。"
        "硬约束："
        "1) 每个arena至少minScripts条；"
        "2) 只可使用该arena提供的groups id，不可发明新id；"
        "3) 每个arena Boss恰好出现1次；且必须在该arena最后一条脚本的后半段（不要求最后一个）；"
        "4) 赔率节奏：前期低赔率占优，中期均衡，后期高赔率占优；"
        "5) gapTimeMs要随阶段逐步收紧，整体呈加压感；"
        "6) scriptId从startScriptId开始全局递增且唯一；type规则：若该行含Boss组则type=2，否则type=1。"
        "风格要求：不要机械重复同一组，避免连续多条都高赔率轰炸，保证可玩性。"
        "输出格式要求：只输出一个合法JSON对象；不要markdown，不要代码块，不要注释。"
        "输出只允许一个JSON对象，结构："
        "{\"scripts\":[{\"scriptId\":123,\"gapTimeMs\":2600,\"arenaIds\":[1],\"type\":1,\"groupIds\":[1,2,3]}],\"notes\":\"一句话说明策略\"}"
    )
    body: Dict[str, Any] = {
        "model": model,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "stream": False,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "task": "根据输入context生成脚本",
                        "context": context,
                    },
                    ensure_ascii=False,
                ),
            },
        ],
    }
    label = f" [{arena_tag}]" if arena_tag else ""
    _push_debug(debug_trace, f"AI endpoint{label}", endpoint)
    _push_debug(debug_trace, f"AI system prompt{label}", system_prompt)

    retry_body = dict(body)
    retry_body.pop("response_format", None)
    retry_body["messages"] = list(body["messages"])
    repair_body = dict(retry_body)
    repair_body["messages"] = list(retry_body["messages"]) + [
        {
            "role": "user",
            "content": "上一条输出未形成完整合法JSON。请重新输出完整JSON对象，仅输出JSON，不要解释。",
        }
    ]

    attempts: List[Tuple[str, Dict[str, Any]]] = [
        ("first-attempt", body),
        ("retry-without-response_format", retry_body),
        ("retry-json-repair", repair_body),
    ]
    last_error = "大模型生成失败"
    for tag, req_body in attempts:
        _push_debug(debug_trace, f"AI request payload ({tag}){label}", req_body)
        try:
            payload = _post_chat_completion(endpoint, api_key, req_body, timeout_sec, debug_trace, f"{tag}{label}")
        except Exception as exc:  # pylint: disable=broad-except
            last_error = str(exc)
            _push_debug(debug_trace, f"AI attempt exception ({tag}){label}", str(exc))
            continue
        finish_reason = _extract_finish_reason(payload)
        if finish_reason:
            _push_debug(debug_trace, f"AI finish_reason ({tag}){label}", finish_reason)

        candidate, source = _extract_completion_candidate(payload)
        _push_debug(debug_trace, f"AI extracted source ({tag}){label}", source)
        if not candidate:
            last_error = f"大模型未返回文本内容（source={source}）"
            continue

        parsed: Dict[str, Any] | None = None
        if isinstance(candidate, dict):
            parsed = candidate
            _push_debug(debug_trace, f"AI extracted candidate dict (source={source}){label}", candidate)
        else:
            content_text = _strip_json_fence(str(candidate))
            _push_debug(debug_trace, f"AI extracted candidate text (source={source}){label}", str(candidate))
            try:
                parsed = json.loads(content_text)
            except json.JSONDecodeError as exc:
                obj_text = _extract_first_json_object(content_text)
                if obj_text:
                    _push_debug(debug_trace, f"AI fallback extracted json object text{label}", obj_text)
                    try:
                        parsed = json.loads(obj_text)
                    except json.JSONDecodeError as exc2:
                        parsed = None
                        last_error = f"大模型返回不是合法JSON: {exc2}（source={source}）"
                else:
                    if finish_reason == "length":
                        last_error = f"大模型输出被截断（finish_reason=length，max_tokens={max_tokens}）"
                    else:
                        last_error = f"大模型返回不是合法JSON: {exc}（source={source}）"
                    parsed = None

        if not isinstance(parsed, dict):
            continue

        scripts = parsed.get("scripts", [])
        if not isinstance(scripts, list):
            last_error = "大模型JSON缺少scripts数组"
            continue
        notes = str(parsed.get("notes", "") or "")
        return scripts, model, notes

    raise RuntimeError(last_error)


def validate_ai_scripts(scripts: List[Dict[str, Any]], context: Dict[str, Any]) -> List[Dict[str, Any]]:
    min_per_arena = max(1, to_int(context.get("minPerArena"), 6))
    start_script_id = max(1, to_int(context.get("startScriptId"), 1))

    allowed_by_arena: Dict[int, set[int]] = {}
    boss_by_arena: Dict[int, set[int]] = {}
    for arena in context.get("arenas", []):
        arena_id = to_int(arena.get("arenaId"), 0)
        groups = arena.get("groups", [])
        allowed = {to_int(g.get("id"), 0) for g in groups if to_int(g.get("id"), 0) > 0}
        bosses = {to_int(g.get("id"), 0) for g in groups if bool(g.get("hasBoss")) and to_int(g.get("id"), 0) > 0}
        if arena_id > 0:
            allowed_by_arena[arena_id] = allowed
            boss_by_arena[arena_id] = bosses

    normalized, warnings = normalize_scripts_payload(scripts)
    if warnings and len(normalized) < 3:
        raise ValueError("大模型生成结果有效行过少")
    if not normalized:
        raise ValueError("大模型未生成有效脚本")

    rows_by_arena: Dict[int, List[Dict[str, Any]]] = {}
    for row in normalized:
        arena_ids = parse_id_list(row.get("arenaIds"))
        if len(arena_ids) != 1:
            raise ValueError("大模型结果要求每行仅一个arena")
        arena_id = arena_ids[0]
        if arena_id not in allowed_by_arena:
            raise ValueError(f"大模型返回了未知arena: {arena_id}")
        filtered_groups = [gid for gid in parse_id_list(row.get("groupIds")) if gid in allowed_by_arena[arena_id]]
        if not filtered_groups:
            non_boss_candidates = sorted([gid for gid in allowed_by_arena[arena_id] if gid not in boss_by_arena.get(arena_id, set())])
            fallback_gid = non_boss_candidates[0] if non_boss_candidates else 0
            if fallback_gid <= 0:
                raise ValueError(f"Arena {arena_id} 行未包含可用group")
            filtered_groups = [fallback_gid]
        row["arenaIds"] = [arena_id]
        row["groupIds"] = filtered_groups
        boss_candidates = boss_by_arena.get(arena_id, set())
        row["type"] = 2 if any(gid in boss_candidates for gid in filtered_groups) else 1
        row["gapTimeMs"] = max(0, to_int(row.get("gapTimeMs"), 0))
        rows_by_arena.setdefault(arena_id, []).append(row)

    for arena_id, rows in rows_by_arena.items():
        if len(rows) < min_per_arena:
            raise ValueError(f"Arena {arena_id} 生成脚本不足{min_per_arena}条")
        boss_candidates = boss_by_arena.get(arena_id, set())
        non_boss_pool = sorted([gid for gid in allowed_by_arena.get(arena_id, set()) if gid not in boss_candidates])
        fallback_gid = non_boss_pool[0] if non_boss_pool else 0

        def scan_boss() -> List[Tuple[int, int, int]]:
            positions: List[Tuple[int, int, int]] = []
            for idx, row in enumerate(rows):
                for pos, gid in enumerate(row["groupIds"]):
                    if gid in boss_candidates:
                        positions.append((idx, pos, gid))
            return positions

        boss_positions = scan_boss()
        if not boss_positions:
            raise ValueError(f"Arena {arena_id} Boss出现次数必须为1，当前为0")

        if len(boss_positions) > 1:
            # Keep the last boss hit, replace/remove all earlier boss hits.
            keep = boss_positions[-1]
            for idx, pos, _gid in reversed(boss_positions[:-1]):
                gids = rows[idx]["groupIds"]
                if 0 <= pos < len(gids):
                    if fallback_gid > 0:
                        gids[pos] = fallback_gid
                    else:
                        gids.pop(pos)
                if not gids:
                    if fallback_gid <= 0:
                        raise ValueError(f"Arena {arena_id} 缺少可替换普通组，无法去除重复Boss")
                    gids.append(fallback_gid)
            boss_positions = scan_boss()
            if len(boss_positions) != 1:
                raise ValueError(f"Arena {arena_id} Boss出现次数必须为1，当前为{len(boss_positions)}")

        boss_row_index, boss_pos, boss_gid = boss_positions[0]
        target_row = rows[-1]
        target_need_pos = max(1, len(target_row["groupIds"]) // 2)
        boss_not_in_final_half = boss_row_index != len(rows) - 1 or boss_pos < target_need_pos
        if boss_not_in_final_half:
            # Remove boss from source row first.
            src = rows[boss_row_index]["groupIds"]
            removed = False
            for i in range(len(src) - 1, -1, -1):
                if src[i] == boss_gid:
                    src.pop(i)
                    removed = True
                    break
            if not removed:
                raise ValueError(f"Arena {arena_id} Boss定位失败")
            if not src:
                if fallback_gid <= 0:
                    raise ValueError(f"Arena {arena_id} 缺少可替换普通组，无法移动Boss")
                src.append(fallback_gid)

            # Ensure final row contains exactly one boss and place it in latter half.
            dst = [gid for gid in rows[-1]["groupIds"] if gid != boss_gid]
            insert_pos = max(1, (len(dst) + 1) // 2)
            if insert_pos > len(dst):
                insert_pos = len(dst)
            dst.insert(insert_pos, boss_gid)
            rows[-1]["groupIds"] = dst

        # Final hard-check: exactly one boss and in last script latter half.
        boss_positions = scan_boss()
        if len(boss_positions) != 1:
            raise ValueError(f"Arena {arena_id} Boss出现次数必须为1，当前为{len(boss_positions)}")
        boss_row_index, boss_pos, _boss_gid = boss_positions[0]
        if boss_row_index != len(rows) - 1:
            raise ValueError(f"Arena {arena_id} Boss必须在最后一条脚本中")
        last_row_len = len(rows[boss_row_index]["groupIds"])
        if boss_pos < max(1, last_row_len // 2):
            raise ValueError(f"Arena {arena_id} Boss必须位于最后阶段（后半段）")

        for row in rows:
            row["type"] = 2 if any(gid in boss_candidates for gid in row["groupIds"]) else 1

    # Reassign continuous script ids to avoid conflicts with existing ids.
    final_rows: List[Dict[str, Any]] = []
    sid = start_script_id
    for arena_id in sorted(rows_by_arena.keys()):
        for row in rows_by_arena[arena_id]:
            row["scriptId"] = sid
            sid += 1
            final_rows.append(row)
    return final_rows


def generate_scripts_by_llm(
    min_per_arena: int, llm_config: Dict[str, Any], debug_trace: List[Dict[str, Any]]
) -> Dict[str, Any]:
    cache = load_all_data()
    context = build_llm_generation_context(cache, min_per_arena)
    _push_debug(debug_trace, "AI generation context", context)
    if not context.get("arenas"):
        raise RuntimeError("当前配置无法构造大模型生成上下文")

    arenas = [a for a in context.get("arenas", []) if isinstance(a, dict) and to_int(a.get("arenaId"), 0) > 0]
    if not arenas:
        raise RuntimeError("AI上下文中没有可用场次")

    next_script_id = max(1, to_int(context.get("startScriptId"), 1))
    all_rows: List[Dict[str, Any]] = []
    model_name = str(llm_config.get("model") or "")
    note_parts: List[str] = []

    for arena in arenas:
        arena_id = to_int(arena.get("arenaId"), 0)
        arena_name = str(arena.get("name") or f"Arena-{arena_id}")
        arena_context = {
            "startScriptId": next_script_id,
            "minPerArena": min_per_arena,
            "arenas": [arena],
        }
        _push_debug(debug_trace, f"AI arena context[{arena_id}]", arena_context)
        scripts_raw, model, notes = call_openai_for_scripts(
            arena_context, llm_config, debug_trace, arena_tag=f"arena:{arena_id}"
        )
        model_name = model or model_name
        _push_debug(debug_trace, f"AI parsed scripts before validation[{arena_id}]", scripts_raw)
        validated_rows = validate_ai_scripts(scripts_raw, arena_context)
        _push_debug(debug_trace, f"AI scripts after validation[{arena_id}]", validated_rows)
        if validated_rows:
            next_script_id = max(to_int(validated_rows[-1].get("scriptId"), next_script_id), next_script_id) + 1
            all_rows.extend(validated_rows)
            if notes:
                note_parts.append(f"{arena_name}: {notes}")

    if not all_rows:
        raise RuntimeError("大模型未生成任何有效脚本")
    return {"scripts": all_rows, "model": model_name, "notes": " | ".join(note_parts)}


def write_script_table(scripts: List[Dict[str, Any]], output_name: str) -> Dict[str, Any]:
    output_path = build_output_path(output_name)
    wb = load_workbook(SCRIPT_FILE)
    ws = wb.worksheets[0]
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    normalized: List[Dict[str, Any]] = []
    warnings: List[str] = []
    for idx, row in enumerate(scripts):
        script_id = to_int(row.get("scriptId"))
        group_ids = parse_id_list(row.get("groupIds"))
        arena_ids = parse_id_list(row.get("arenaIds"))
        gap_time = to_int(row.get("gapTimeMs"), 0)
        row_type = to_int(row.get("type"), 1)
        if script_id <= 0:
            warnings.append(f"第 {idx + 1} 行 scriptId 非法，已跳过")
            continue
        if not group_ids:
            warnings.append(f"第 {idx + 1} 行 groupIds 为空，已跳过")
            continue
        if not arena_ids:
            warnings.append(f"第 {idx + 1} 行 arenaIds 为空，已跳过")
            continue
        normalized.append(
            {
                "scriptId": script_id,
                "groupIds": group_ids,
                "gapTimeMs": max(0, gap_time),
                "arenaIds": arena_ids,
                "type": row_type,
            }
        )

    seen_ids = set()
    for row in normalized:
        sid = row["scriptId"]
        if sid in seen_ids:
            warnings.append(f"ScriptId={sid} 重复，运行逻辑可能冲突")
        seen_ids.add(sid)

    for index, row in enumerate(normalized, start=3):
        ws.cell(index, 1, row["scriptId"])
        ws.cell(index, 2, ",".join(str(x) for x in row["groupIds"]))
        ws.cell(index, 3, row["gapTimeMs"])
        ws.cell(index, 4, ",".join(str(x) for x in row["arenaIds"]))
        ws.cell(index, 5, row["type"])

    wb.save(output_path)
    return {
        "savedFile": output_path.name,
        "savedPath": str(output_path),
        "rows": len(normalized),
        "baseFile": SCRIPT_FILE.name,
        "basePreserved": True,
        "warnings": warnings,
    }


def guess_mime(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".html":
        return "text/html; charset=utf-8"
    if suffix == ".js":
        return "application/javascript; charset=utf-8"
    if suffix == ".css":
        return "text/css; charset=utf-8"
    if suffix == ".json":
        return "application/json; charset=utf-8"
    return "application/octet-stream"


class EditorHandler(BaseHTTPRequestHandler):
    def _write_json(self, data: Any, status: int = 200) -> None:
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def _write_text(self, text: str, status: int = 200) -> None:
        payload = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def _serve_static(self, rel_path: str) -> None:
        rel = rel_path.strip("/") or "index.html"
        clean = Path(rel)
        if clean.is_absolute() or ".." in clean.parts:
            self._write_text("invalid path", 400)
            return
        file_path = WEB_DIR / clean
        if not file_path.exists() or not file_path.is_file():
            self._write_text("not found", 404)
            return
        data = file_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", guess_mime(file_path))
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/data":
            cache = load_all_data()
            self._write_json(
                {
                    "ok": True,
                    "tableDir": str(TABLE_DIR),
                    "logicHint": {
                        "arena": "渔场决定鱼出现在哪个房间",
                        "fish": "基础鱼表决定有哪些鱼（中文名 + 赔率）",
                        "group": "鱼按组合配置（Group）",
                        "route": "鱼沿路径出现（Route）",
                        "script": "脚本配置组合 + 时间节奏（GapTime）",
                    },
                    "arenas": cache.arenas,
                    "fish": cache.fish,
                    "groups": cache.groups,
                    "routes": cache.routes,
                    "scripts": cache.scripts,
                }
            )
            return

        if parsed.path == "/api/presets":
            self._write_json({"ok": True, "presets": list_presets()})
            return

        if parsed.path == "/api/health":
            self._write_json({"ok": True})
            return

        if parsed.path == "/" or parsed.path.startswith("/index.html"):
            self._serve_static("index.html")
            return
        self._serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path not in (
            "/api/save-script",
            "/api/preset/save",
            "/api/preset/load",
            "/api/preset/delete",
            "/api/generate-script-ai",
        ):
            self._write_text("not found", 404)
            return

        length = to_int(self.headers.get("Content-Length"), 0)
        body = self.rfile.read(length) if length > 0 else b"{}"
        try:
            payload = json.loads(body.decode("utf-8"))
        except json.JSONDecodeError:
            self._write_json({"ok": False, "error": "json parse error"}, 400)
            return

        if parsed.path == "/api/generate-script-ai":
            debug_trace: List[Dict[str, Any]] = []
            try:
                min_per_arena = max(1, to_int(payload.get("minPerArena"), 6))
                llm_config = sanitize_llm_config(payload.get("llmConfig"))
                debug_cfg = dict(llm_config)
                if debug_cfg.get("apiKey"):
                    api_key = str(debug_cfg.get("apiKey"))
                    debug_cfg["apiKey"] = f"***{api_key[-4:]}" if len(api_key) >= 4 else "***"
                _push_debug(debug_trace, "received llmConfig (masked)", debug_cfg)
                result = generate_scripts_by_llm(min_per_arena, llm_config, debug_trace)
                _debug_print_json("AI debug trace", debug_trace)
                self._write_json({"ok": True, **result, "debug": debug_trace})
            except Exception as exc:  # pylint: disable=broad-except
                _push_debug(debug_trace, "exception", str(exc))
                _debug_print_json("AI debug trace", debug_trace)
                self._write_json({"ok": False, "error": str(exc), "debug": debug_trace}, 400)
            return

        if parsed.path == "/api/preset/load":
            try:
                name = str(payload.get("name", ""))
                result = load_preset(name)
                self._write_json({"ok": True, **result})
            except Exception as exc:  # pylint: disable=broad-except
                self._write_json({"ok": False, "error": str(exc)}, 400)
            return

        if parsed.path == "/api/preset/delete":
            try:
                name = str(payload.get("name", ""))
                result = delete_preset(name)
                self._write_json({"ok": True, **result})
            except Exception as exc:  # pylint: disable=broad-except
                self._write_json({"ok": False, "error": str(exc)}, 400)
            return

        if parsed.path == "/api/preset/save":
            scripts = payload.get("scripts", [])
            if not isinstance(scripts, list):
                self._write_json({"ok": False, "error": "scripts must be array"}, 400)
                return
            try:
                name = str(payload.get("name", ""))
                meta = payload.get("meta", {})
                if not isinstance(meta, dict):
                    meta = {}
                result = save_preset(name, scripts, meta)
                self._write_json({"ok": True, **result})
            except Exception as exc:  # pylint: disable=broad-except
                self._write_json({"ok": False, "error": str(exc)}, 400)
            return

        scripts = payload.get("scripts", [])
        if not isinstance(scripts, list):
            self._write_json({"ok": False, "error": "scripts must be array"}, 400)
            return

        output_name = str(payload.get("outputFile", "") or "").strip()
        if output_name and not OUTPUT_FILE_RE.match(output_name):
            self._write_json({"ok": False, "error": "outputFile 名称非法"}, 400)
            return

        try:
            result = write_script_table(scripts, output_name)
            self._write_json({"ok": True, **result})
        except Exception as exc:  # pylint: disable=broad-except
            self._write_json({"ok": False, "error": str(exc)}, 500)

    def log_message(self, format: str, *args: Any) -> None:  # noqa: A003
        print("[editor]", format % args)


def run_server(host: str, port: int) -> None:
    if not WEB_DIR.exists():
        raise RuntimeError(f"缺少前端目录: {WEB_DIR}")
    server = ThreadingHTTPServer((host, port), EditorHandler)
    print(f"Fish Table Editor running: http://{host}:{port}")
    print(f"Table dir: {TABLE_DIR}")
    print("Press Ctrl+C to stop")
    server.serve_forever()


def run_check() -> None:
    cache = load_all_data()
    print("check ok")
    print(f"arenas: {len(cache.arenas)}")
    print(f"fish: {len(cache.fish)}")
    print(f"groups: {len(cache.groups)}")
    print(f"routes: {len(cache.routes)}")
    print(f"scripts: {len(cache.scripts)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Fish table visual editor server")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=18888)
    parser.add_argument("--check", action="store_true", help="only check table parsing")
    args = parser.parse_args()

    if args.check:
        run_check()
        return
    run_server(args.host, args.port)


if __name__ == "__main__":
    main()

